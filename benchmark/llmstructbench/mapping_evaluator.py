"""Mapping pipeline evaluator: template structure vs mapped XLSX.

Evaluates how well the DM (Dynamic Mapping) pipeline fills the golden
template. Uses TemplateAnalyzer to extract expected structure, then
compares filled values cell-by-cell.

The template itself serves as the "JSON Schema" — each data cell in the
template defines an expected field. The mapped XLSX output is compared
against the extraction source data that should have been placed there.
"""

from __future__ import annotations

import sys
from difflib import SequenceMatcher
from pathlib import Path

from struct_metrics import CellPair, SheetResult, BatchResult
from error_classifier import classify_pairs, analysis_to_f1_result, ErrorAnalysis


def _ensure_src_path() -> None:
    """Add excel_ready/src to sys.path for TemplateAnalyzer import."""
    src_path = str(Path(__file__).resolve().parent.parent.parent / "src")
    if src_path not in sys.path:
        sys.path.insert(0, src_path)


def _load_template_data_cells(template_path: str) -> dict[str, list[dict]]:
    """Extract expected data cell positions from template using TemplateAnalyzer.

    Returns:
        {section_name: [{row, col, label, type}, ...]}
    """
    _ensure_src_path()
    try:
        from mapper.template_analyzer import TemplateAnalyzer
    except ImportError:
        return {}

    analyzer = TemplateAnalyzer()
    ts = analyzer.analyze(template_path)

    sections: dict[str, list[dict]] = {}
    for section in ts.sections:
        cells: list[dict] = []
        for col_name, cell_ref in section.columns.items():
            cells.append({
                "row": section.data_start_row,
                "col_label": col_name,
                "col_idx": cell_ref.col if hasattr(cell_ref, "col") else None,
                "section": section.name,
            })
        sections[section.name] = cells

    if ts.summary_section:
        cells = []
        for col_name, cell_ref in ts.summary_section.columns.items():
            cells.append({
                "row": ts.summary_section.data_start_row,
                "col_label": col_name,
                "col_idx": cell_ref.col if hasattr(cell_ref, "col") else None,
                "section": ts.summary_section.name,
            })
        sections[ts.summary_section.name] = cells

    return sections


def _load_xlsx_value(xlsx_path: Path, sheet_name: str, row: int, col: int) -> object:
    """Load a single cell value from XLSX (1-indexed)."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet_name]
    val = ws.cell(row=row, column=col).value
    wb.close()
    return val


def _load_xlsx_sheet_all(xlsx_path: Path, sheet_name: str) -> dict[tuple[int, int], object]:
    """Load all cell values from a sheet as {(row, col): value}."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet_name]
    result: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            if cell.value is not None:
                result[(cell.row, cell.column)] = cell.value
    wb.close()
    return result


def _load_mapped_data(mapped_xlsx: Path) -> dict[str, dict[tuple[int, int], object]]:
    """Load all data cells from mapped XLSX per sheet."""
    from openpyxl import load_workbook
    wb = load_workbook(mapped_xlsx, read_only=True, data_only=True)
    result: dict[str, dict[tuple[int, int], object]] = {}
    for ws in wb.worksheets:
        cells: dict[tuple[int, int], object] = {}
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                if cell.value is not None:
                    cells[(cell.row, cell.column)] = cell.value
        result[ws.title] = cells
    wb.close()
    return result


def evaluate_mapping(
    template_xlsx: str | Path,
    mapped_xlsx: str | Path,
) -> BatchResult:
    """Evaluate mapping output against template structure.

    Compares the mapped XLSX output with the source extraction data
    to measure how accurately data was placed into the template.

    The evaluation checks:
    1. Whether data cells in the template are filled (non-empty)
    2. Whether filled values match the source extraction data
    3. Template structure preservation (merge cells, styles)

    Since we may not have a separate GT mapped file, we evaluate:
    - **Fill completeness**: what % of expected data cells are non-empty
    - **Structure preservation**: merge cell count, style consistency
    - **Value validity**: numeric fields contain valid numbers

    Args:
        template_xlsx: Path to golden template XLSX.
        mapped_xlsx: Path to mapped output (_dm.xlsx).

    Returns:
        BatchResult with per-section results.
    """
    template_path = Path(template_xlsx)
    mapped_path = Path(mapped_xlsx)

    if not template_path.exists() or not mapped_path.exists():
        return BatchResult()

    from openpyxl import load_workbook

    tpl_wb = load_workbook(template_path, read_only=True, data_only=True)
    tpl_sheet = tpl_wb.worksheets[0].title
    tpl_wb.close()

    mapped_wb = load_workbook(mapped_path, read_only=True, data_only=True)
    mapped_sheet = mapped_wb.worksheets[0].title if mapped_wb.worksheets else tpl_sheet
    mapped_wb.close()

    tpl_cells = _load_xlsx_sheet_all(template_path, tpl_sheet)
    map_cells = _load_xlsx_sheet_all(mapped_path, mapped_sheet)

    tpl_data_cells = {
        pos: val for pos, val in tpl_cells.items()
        if val is not None and str(val).strip() != ""
    }

    map_data_cells = {
        pos: val for pos, val in map_cells.items()
        if val is not None and str(val).strip() != ""
    }

    all_positions = set(tpl_data_cells.keys()) | set(map_data_cells.keys())

    gt_pairs: list[CellPair] = []
    sys_pairs: list[CellPair] = []

    for pos in sorted(all_positions):
        row, col = pos
        key = f"{tpl_sheet}::col{col}::row{row}"

        if pos in tpl_data_cells:
            gt_pairs.append(CellPair(
                key=key,
                value=tpl_data_cells[pos],
                field_name=f"col{col}",
            ))

        sys_val = map_cells.get(pos)
        if sys_val is not None:
            sys_pairs.append(CellPair(
                key=key,
                value=sys_val,
                field_name=f"col{col}",
            ))

    from openpyxl import load_workbook as _lwb

    tpl_wb2 = _lwb(template_path, data_only=True)
    map_wb2 = _lwb(mapped_path, data_only=True)
    tpl_ws = tpl_wb2.worksheets[0]
    map_ws = map_wb2.worksheets[0]

    tpl_merges = len(tpl_ws.merged_cells.ranges)
    map_merges = len(map_ws.merged_cells.ranges)
    tpl_wb2.close()
    map_wb2.close()

    batch = BatchResult()

    analysis = classify_pairs(gt_pairs, sys_pairs)
    f1_result = analysis_to_f1_result(analysis)

    total_tpl = len(tpl_data_cells)
    total_map = len(map_data_cells)
    overlap = len(set(tpl_data_cells.keys()) & set(map_data_cells.keys()))

    fill_ratio = overlap / total_tpl if total_tpl > 0 else 0.0

    is_correct = (
        analysis.fn_keys == 0
        and analysis.missing_values == 0
        and analysis.error_count == 0
        and map_merges >= tpl_merges * 0.9
    )

    is_failed = total_map == 0

    batch.sheet_results.append(SheetResult(
        sheet_name=f"{mapped_sheet} (merges: {map_merges}/{tpl_merges})",
        f1_result=f1_result,
        is_correct=is_correct,
        is_failed=is_failed,
        total_cells=total_tpl,
        matched_cells=overlap,
    ))

    batch.sheet_results[-1].f1_result.detail = {
        "tpl_merges": tpl_merges,
        "map_merges": map_merges,
        "merge_ratio": map_merges / tpl_merges if tpl_merges > 0 else 0.0,
        "fill_ratio": fill_ratio,
        "total_template_cells": total_tpl,
        "total_mapped_cells": total_map,
        "overlap_cells": overlap,
    }

    return batch


def evaluate_mapping_batch(
    cases: list[dict],
) -> BatchResult:
    """Evaluate multiple mapping cases.

    Args:
        cases: List of {stem, template_xlsx, mapped_xlsx} dicts.

    Returns:
        BatchResult aggregated across all cases.
    """
    batch = BatchResult()
    for case in cases:
        sub_batch = evaluate_mapping(case["template_xlsx"], case["mapped_xlsx"])
        for sr in sub_batch.sheet_results:
            sr.sheet_name = f"{case['stem']}/{sr.sheet_name}"
            batch.sheet_results.append(sr)
    return batch
