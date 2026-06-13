"""Extraction pipeline evaluator: GT XLSX vs system extraction XLSX.

Loads ground-truth and system output XLSX files, extracts cell pairs
per sheet, and evaluates using LLMStructBench metrics.

Supports fuzzy header matching when column names differ slightly.
"""

from __future__ import annotations

from difflib import SequenceMatcher
from pathlib import Path

from struct_metrics import CellPair, SheetResult, BatchResult
from error_classifier import classify_pairs, analysis_to_f1_result, ErrorAnalysis


# ════════════════════════════════════════════════════════════
# XLSX loading
# ════════════════════════════════════════════════════════════

def _load_xlsx_sheet_pairs(
    xlsx_path: Path,
    sheet_name: str,
    max_rows: int = 500,
) -> tuple[list[str], list[list[object]]]:
    """Load a single sheet's headers and data rows.

    Returns:
        (headers, rows) where headers is a list of column names
        and rows is a list of row-value-lists.
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return [], []

    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    headers = [str(h).strip() if h else "" for h in header_row]

    rows: list[list[object]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i >= max_rows:
            break
        rows.append(list(row))

    wb.close()
    return headers, rows


def _match_headers(
    gt_headers: list[str],
    sys_headers: list[str],
    threshold: float = 0.85,
) -> dict[int, int]:
    """Match GT headers to system headers by index and fuzzy name.

    Returns:
        {gt_col_index: sys_col_index}
    """
    result: dict[int, int] = {}

    sys_used: set[int] = set()
    for gt_i, gt_h in enumerate(gt_headers):
        if not gt_h:
            continue

        if gt_i < len(sys_headers) and sys_headers[gt_i] == gt_h:
            result[gt_i] = gt_i
            sys_used.add(gt_i)
            continue

        best_j = -1
        best_ratio = 0.0
        for sys_j, sys_h in enumerate(sys_headers):
            if sys_j in sys_used or not sys_h:
                continue
            ratio = SequenceMatcher(None, gt_h.lower(), sys_h.lower()).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_j = sys_j

        if best_j >= 0 and best_ratio >= threshold:
            result[gt_i] = best_j
            sys_used.add(best_j)

    return result


def _build_cell_pairs(
    headers: list[str],
    rows: list[list[object]],
    sheet_name: str,
) -> list[CellPair]:
    """Build CellPair list from headers and rows.

    Key format: "{sheet_name}::{header}::{row_num}"
    """
    pairs: list[CellPair] = []
    for row_idx, row in enumerate(rows):
        for col_idx, val in enumerate(row):
            if col_idx >= len(headers) or not headers[col_idx]:
                continue
            header = headers[col_idx]
            pairs.append(CellPair(
                key=f"{sheet_name}::{header}::R{row_idx + 1:04d}",
                value=val,
                field_name=header,
            ))
    return pairs


# ════════════════════════════════════════════════════════════
# Evaluation
# ════════════════════════════════════════════════════════════

def evaluate_extraction(
    gt_xlsx: str | Path,
    sys_xlsx: str | Path,
) -> BatchResult:
    """Evaluate extraction output against ground truth.

    Args:
        gt_xlsx: Path to ground-truth XLSX (right_files/).
        sys_xlsx: Path to system extraction output (_맵핑.xlsx).

    Returns:
        BatchResult with per-sheet results.
    """
    gt_path = Path(gt_xlsx)
    sys_path = Path(sys_xlsx)

    if not gt_path.exists():
        return BatchResult()
    if not sys_path.exists():
        return BatchResult()

    from openpyxl import load_workbook

    gt_wb = load_workbook(gt_path, read_only=True, data_only=True)
    gt_sheet_names = [ws.title for ws in gt_wb.worksheets]
    gt_wb.close()

    sys_wb = load_workbook(sys_path, read_only=True, data_only=True)
    sys_sheet_names = [ws.title for ws in sys_wb.worksheets]
    sys_wb.close()

    batch = BatchResult()

    for sheet_name in gt_sheet_names:
        gt_headers, gt_rows = _load_xlsx_sheet_pairs(gt_path, sheet_name)
        sys_headers, sys_rows = _load_xlsx_sheet_pairs(sys_path, sheet_name)

        if not gt_headers:
            continue

        if not sys_headers and sheet_name not in sys_sheet_names:
            batch.sheet_results.append(SheetResult(
                sheet_name=sheet_name,
                f1_result=analysis_to_f1_result(ErrorAnalysis(
                    tp_keys=0, fp_keys=0, fn_keys=len(gt_headers),
                )),
                is_correct=False,
                is_failed=True,
                total_cells=0,
                matched_cells=0,
            ))
            continue

        header_map = _match_headers(gt_headers, sys_headers)

        gt_pairs: list[CellPair] = []
        sys_pairs: list[CellPair] = []

        for gt_col, sys_col in header_map.items():
            gt_h = gt_headers[gt_col]
            for row_idx in range(max(len(gt_rows), len(sys_rows))):
                gt_val = gt_rows[row_idx][gt_col] if row_idx < len(gt_rows) and gt_col < len(gt_rows[row_idx]) else None
                sys_val = sys_rows[row_idx][sys_col] if row_idx < len(sys_rows) and sys_col < len(sys_rows[row_idx]) else None

                key = f"{sheet_name}::{gt_h}::R{row_idx + 1:04d}"
                gt_pairs.append(CellPair(key=key, value=gt_val, field_name=gt_h))
                sys_pairs.append(CellPair(key=key, value=sys_val, field_name=gt_h))

        for gt_col in range(len(gt_headers)):
            if gt_col not in header_map and gt_headers[gt_col]:
                for row_idx in range(len(gt_rows)):
                    gt_val = gt_rows[row_idx][gt_col] if gt_col < len(gt_rows[row_idx]) else None
                    key = f"{sheet_name}::{gt_headers[gt_col]}::R{row_idx + 1:04d}"
                    gt_pairs.append(CellPair(key=key, value=gt_val, field_name=gt_headers[gt_col]))

        for sys_col in range(len(sys_headers)):
            if sys_col not in set(header_map.values()) and sys_headers[sys_col]:
                for row_idx in range(len(sys_rows)):
                    sys_val = sys_rows[row_idx][sys_col] if sys_col < len(sys_rows[row_idx]) else None
                    key = f"{sheet_name}::{sys_headers[sys_col]}::R{row_idx + 1:04d}"
                    sys_pairs.append(CellPair(key=key, value=sys_val, field_name=sys_headers[sys_col]))

        analysis = classify_pairs(gt_pairs, sys_pairs)
        f1_result = analysis_to_f1_result(analysis)

        is_correct = analysis.fn_keys == 0 and analysis.missing_values == 0 and analysis.error_count == 0

        batch.sheet_results.append(SheetResult(
            sheet_name=sheet_name,
            f1_result=f1_result,
            is_correct=is_correct,
            is_failed=False,
            total_cells=len(gt_pairs),
            matched_cells=int(analysis.tp_values),
        ))

    return batch


def evaluate_extraction_batch(
    cases: list[dict],
) -> BatchResult:
    """Evaluate multiple extraction cases.

    Args:
        cases: List of {stem, gt_xlsx, sys_xlsx} dicts.

    Returns:
        BatchResult aggregated across all cases.
    """
    batch = BatchResult()
    for case in cases:
        sub_batch = evaluate_extraction(case["gt_xlsx"], case["sys_xlsx"])
        for sr in sub_batch.sheet_results:
            sr.sheet_name = f"{case['stem']}/{sr.sheet_name}"
            batch.sheet_results.append(sr)
    return batch
