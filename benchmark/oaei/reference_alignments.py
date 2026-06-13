"""Load reference alignments from mapping_rules_store.json and extraction data.

Reference alignments serve as the gold standard for OAEI evaluation.
Two layers are supported:

1. **Mapping-layer reference**: User-confirmed field mappings from
   `mapping_rules_store.json` (source_pattern <-> target_field per section).
   These are human-validated alignments between extraction fields and
   template fields.

2. **Extraction-layer reference**: Ground-truth XLSX headers from
   `right_files/` compared against pipeline output headers.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from oaei_metrics import Alignment


def load_mapping_reference(
    store_path: str | Path,
    template_name_filter: str | None = None,
) -> dict[str, list[Alignment]]:
    """Load user-confirmed mappings as reference alignments.

    Args:
        store_path: Path to mapping_rules_store.json.
        template_name_filter: If set, only include rules for this template.

    Returns:
        {case_stem: [Alignment, ...]}
    """
    store_path = Path(store_path)
    if not store_path.exists():
        print(f"[WARN] mapping_rules_store not found: {store_path}")
        return {}

    raw = json.loads(store_path.read_text("utf-8"))

    if isinstance(raw, dict) and "rules" in raw:
        data = raw["rules"]
    elif isinstance(raw, list):
        data = raw
    else:
        data = []

    per_case: dict[str, list[Alignment]] = {}
    for rule in data:
        if not isinstance(rule, dict):
            continue
        if template_name_filter and rule.get("template_name") != template_name_filter:
            continue

        case_stem = rule.get("case_stem", "unknown")
        source = rule.get("source_pattern", "") or rule.get("source_field", "")
        target = rule.get("target_field", "")
        section = rule.get("section", "")
        confidence = rule.get("confidence", 1.0)

        if source and target:
            key = f"{section}::{source}" if section else source
            val = f"{section}::{target}" if section else target
            per_case.setdefault(case_stem, []).append(
                Alignment(source=key, target=val, confidence=confidence)
            )

    return per_case


def load_extraction_reference(
    right_files_dir: str | Path,
) -> dict[str, list[Alignment]]:
    """Load ground-truth extraction headers as reference alignments.

    Uses the column headers from ground-truth XLSX files as the
    reference set. Each header is both source and target (identity
    alignment) since extraction should reproduce the exact headers.

    Args:
        right_files_dir: Directory containing ground-truth XLSX files.

    Returns:
        {case_stem: [Alignment, ...]}
    """
    right_files_dir = Path(right_files_dir)
    if not right_files_dir.exists():
        print(f"[WARN] right_files dir not found: {right_files_dir}")
        return {}

    per_case: dict[str, list[Alignment]] = {}

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[WARN] openpyxl not available, skipping extraction reference")
        return {}

    for xlsx_path in sorted(right_files_dir.glob("*.xlsx")):
        stem = xlsx_path.stem
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                sheet_name = ws.title
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header_row:
                    continue
                for col_idx, val in enumerate(header_row):
                    if val and str(val).strip():
                        field_name = str(val).strip()
                        per_case.setdefault(stem, []).append(
                            Alignment(
                                source=f"{sheet_name}::{field_name}",
                                target=f"{sheet_name}::{field_name}",
                                confidence=1.0,
                            )
                        )
            wb.close()
        except Exception as exc:
            print(f"[WARN] Failed to read {xlsx_path}: {exc}")

    return per_case


def get_template_columns(template_path: str | Path, sheet_name: str = "") -> dict[str, list[str]]:
    """Extract column labels from template using TemplateAnalyzer.

    Args:
        template_path: Path to template XLSX.
        sheet_name: Optional sheet name.

    Returns:
        {section_name: [column_label, ...]}
    """
    template_path = str(template_path)

    src_path = str(Path(__file__).resolve().parent.parent.parent / "src")
    if src_path not in sys.path:
        sys.path.insert(0, src_path)

    try:
        from mapper.template_analyzer import TemplateAnalyzer
    except ImportError as exc:
        print(f"[ERROR] Cannot import TemplateAnalyzer: {exc}")
        return {}

    analyzer = TemplateAnalyzer()
    ts = analyzer.analyze(template_path, sheet_name)

    sections: dict[str, list[str]] = {}
    for section in ts.sections:
        sections[section.name] = list(section.columns.keys())
    if ts.summary_section:
        sections[ts.summary_section.name] = list(ts.summary_section.columns.keys())

    return sections


def get_extraction_columns(xlsx_path: str | Path) -> dict[str, list[str]]:
    """Extract column headers from pipeline output XLSX.

    Args:
        xlsx_path: Path to _맵핑.xlsx or similar.

    Returns:
        {sheet_name: [header, ...]}
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return {}

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    result: dict[str, list[str]] = {}
    for ws in wb.worksheets:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        headers = [
            str(v).strip()
            for v in header_row
            if v and str(v).strip()
        ]
        result[ws.title] = headers
    wb.close()
    return result
